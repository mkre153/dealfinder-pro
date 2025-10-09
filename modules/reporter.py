from typing import Dict, List, Optional
from jinja2 import Environment, FileSystemLoader, select_autoescape
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
import os

class ReportGenerator:
    """Generates email and Excel reports"""

    def __init__(self, config: Dict):
        self.config = config
        self.logger = logging.getLogger(__name__)

        # Setup Jinja2 for email templates
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
        os.makedirs(template_dir, exist_ok=True)

        self.jinja_env = Environment(
            loader=FileSystemLoader(template_dir),
            autoescape=select_autoescape(['html', 'xml'])
        )

    def generate_daily_email_report(self, properties: List[Dict],
                                    stats: Optional[Dict] = None) -> str:
        """
        Generate HTML email with top properties

        Includes:
        - Summary statistics (total found, avg score, hot deals)
        - Top 10 properties ranked by score
        - Property photos, key stats, analysis
        - Price reductions from yesterday
        - Market insights

        Args:
            properties: List of analyzed property dicts
            stats: Optional stats dict with additional metrics

        Returns:
            HTML string for email body
        """
        try:
            # Sort by score descending
            top_properties = sorted(
                properties,
                key=lambda x: x.get('opportunity_score', 0),
                reverse=True
            )[:10]

            # Calculate summary stats
            total_properties = len(properties)
            hot_deals = len([p for p in properties if p.get('opportunity_score', 0) >= 90])
            good_deals = len([p for p in properties if 75 <= p.get('opportunity_score', 0) < 90])

            avg_score = (sum(p.get('opportunity_score', 0) for p in properties) /
                        len(properties)) if properties else 0

            # Properties with price reductions
            price_reductions = [p for p in properties
                              if p.get('price_reduction_amount', 0) > 0]
            price_reductions.sort(key=lambda x: x.get('price_reduction_amount', 0), reverse=True)

            # Market insights
            total_value = sum(p.get('list_price', 0) for p in properties)
            avg_price = total_value / len(properties) if properties else 0

            if stats is None:
                stats = {}

            template = self.jinja_env.get_template('daily_report.html')

            html = template.render(
                date=datetime.now().strftime('%B %d, %Y'),
                total_properties=total_properties,
                hot_deals=hot_deals,
                good_deals=good_deals,
                avg_score=avg_score,
                top_properties=top_properties,
                price_reductions=price_reductions[:5],
                avg_price=avg_price,
                stats=stats
            )

            self.logger.info(f"Generated daily email report with {len(top_properties)} properties")
            return html

        except Exception as e:
            self.logger.error(f"Error generating daily email report: {e}", exc_info=True)
            return self._generate_error_email(str(e))

    def generate_excel_report(self, properties: List[Dict],
                             filepath: str) -> str:
        """
        Generate Excel report with detailed property data

        Sheets:
        1. Summary - Top deals with key metrics
        2. All Properties - Complete dataset
        3. Market Analysis - ZIP code statistics

        Args:
            properties: List of analyzed property dicts
            filepath: Absolute path to save Excel file

        Returns:
            Filepath of saved Excel file
        """
        try:
            wb = openpyxl.Workbook()

            # Sheet 1: Summary (top 20 deals)
            ws_summary = wb.active
            ws_summary.title = "Top Deals"
            self._create_summary_sheet(ws_summary, properties)

            # Sheet 2: All Properties
            ws_all = wb.create_sheet("All Properties")
            self._create_all_properties_sheet(ws_all, properties)

            # Sheet 3: Market Analysis
            ws_market = wb.create_sheet("Market Analysis")
            self._create_market_analysis_sheet(ws_market, properties)

            # Save
            wb.save(filepath)
            self.logger.info(f"Excel report saved: {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}", exc_info=True)
            raise

    def _create_summary_sheet(self, ws, properties: List[Dict]):
        """Create formatted summary sheet with top deals"""
        # Title
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = f"DealFinder Pro - Top Investment Opportunities ({datetime.now().strftime('%B %d, %Y')})"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Headers
        headers = [
            'Rank', 'Score', 'Deal Quality', 'Address', 'City', 'ZIP',
            'Price', 'Beds/Baths', 'Sqft', 'Below Market %', 'Est. Profit'
        ]
        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        sorted_props = sorted(
            properties,
            key=lambda x: x.get('opportunity_score', 0),
            reverse=True
        )[:20]

        for rank, prop in enumerate(sorted_props, 1):
            row_data = [
                rank,
                prop.get('opportunity_score', 0),
                prop.get('deal_quality', 'N/A'),
                prop.get('street_address', ''),
                prop.get('city', ''),
                prop.get('zip_code', ''),
                prop.get('list_price', 0),
                f"{prop.get('bedrooms', 0)}/{prop.get('bathrooms', 0)}",
                prop.get('square_feet', 0),
                prop.get('below_market_percentage', 0),
                prop.get('estimated_profit', 0)
            ]
            ws.append(row_data)

            # Color code by deal quality
            row_idx = ws.max_row
            deal_quality = prop.get('deal_quality', '')

            if deal_quality == "HOT DEAL":
                fill_color = "FFE6E6"  # Light red
            elif deal_quality == "GOOD OPPORTUNITY":
                fill_color = "E6FFE6"  # Light green
            else:
                fill_color = "FFFFFF"  # White

            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Format currency and percentage columns
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=7, max_col=7):
            for cell in row:
                cell.number_format = '"$"#,##0'

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=10, max_col=10):
            for cell in row:
                cell.number_format = '0.0"%"'

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=11, max_col=11):
            for cell in row:
                cell.number_format = '"$"#,##0'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_all_properties_sheet(self, ws, properties: List[Dict]):
        """Create sheet with all property data"""
        # Headers with comprehensive property details
        headers = [
            'MLS#', 'Address', 'City', 'State', 'ZIP', 'Price',
            'Beds', 'Baths', 'Sqft', 'Price/Sqft', 'Lot Size',
            'Year Built', 'Property Type', 'DOM', 'Score',
            'Deal Quality', 'Below Market %', 'Est. Market Value',
            'Est. Profit', 'Cap Rate', 'Monthly Rent', 'Recommendation'
        ]
        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Add all property data
        for prop in properties:
            metrics = prop.get('investment_metrics', {})

            row_data = [
                prop.get('mls_number', ''),
                prop.get('street_address', ''),
                prop.get('city', ''),
                prop.get('state', ''),
                prop.get('zip_code', ''),
                prop.get('list_price', 0),
                prop.get('bedrooms', 0),
                prop.get('bathrooms', 0),
                prop.get('square_feet', 0),
                metrics.get('price_per_sqft', 0),
                prop.get('lot_size', ''),
                prop.get('year_built', ''),
                prop.get('property_type', ''),
                prop.get('days_on_market', 0),
                prop.get('opportunity_score', 0),
                prop.get('deal_quality', ''),
                prop.get('below_market_percentage', 0),
                prop.get('estimated_market_value', 0),
                prop.get('estimated_profit', 0),
                metrics.get('cap_rate', 0),
                metrics.get('estimated_monthly_rent', 0),
                prop.get('recommendation', '')
            ]
            ws.append(row_data)

        # Format currency columns
        currency_columns = [6, 10, 18, 19, 21]  # Price, Price/Sqft, Market Value, Profit, Rent
        for col in currency_columns:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '"$"#,##0'

        # Format percentage columns
        percentage_columns = [17, 20]  # Below Market %, Cap Rate
        for col in percentage_columns:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '0.0"%"'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_market_analysis_sheet(self, ws, properties: List[Dict]):
        """Create market statistics sheet grouped by ZIP code"""
        # Group properties by ZIP code
        zip_groups = {}

        for prop in properties:
            zip_code = prop.get('zip_code', 'Unknown')
            if zip_code not in zip_groups:
                zip_groups[zip_code] = []
            zip_groups[zip_code].append(prop)

        # Headers
        headers = [
            'ZIP Code', 'Total Properties', 'Avg Score', 'Hot Deals',
            'Avg Price', 'Avg Price/Sqft', 'Avg DOM', 'Avg Below Market %',
            'Total Est. Profit'
        ]
        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Calculate stats for each ZIP code
        for zip_code, props in sorted(zip_groups.items()):
            total = len(props)
            avg_score = sum(p.get('opportunity_score', 0) for p in props) / total
            hot_deals = len([p for p in props if p.get('opportunity_score', 0) >= 90])
            avg_price = sum(p.get('list_price', 0) for p in props) / total

            # Calculate avg price per sqft
            valid_sqft = [(p.get('list_price', 0) / p.get('square_feet', 1))
                         for p in props if p.get('square_feet', 0) > 0]
            avg_price_sqft = sum(valid_sqft) / len(valid_sqft) if valid_sqft else 0

            avg_dom = sum(p.get('days_on_market', 0) for p in props) / total
            avg_below_market = sum(p.get('below_market_percentage', 0) for p in props) / total
            total_profit = sum(p.get('estimated_profit', 0) for p in props)

            row_data = [
                zip_code,
                total,
                avg_score,
                hot_deals,
                avg_price,
                avg_price_sqft,
                avg_dom,
                avg_below_market,
                total_profit
            ]
            ws.append(row_data)

        # Format currency columns
        for col in [5, 6, 9]:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '"$"#,##0'

        # Format percentage column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
            for cell in row:
                cell.number_format = '0.0"%"'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_property_card_html(self, property_data: Dict) -> str:
        """
        Generate HTML card for single property (for email)

        Args:
            property_data: Property details with analysis

        Returns:
            HTML string for property card
        """
        try:
            template = self.jinja_env.get_template('property_card.html')
            return template.render(property=property_data)
        except Exception as e:
            self.logger.error(f"Error generating property card: {e}")
            return f"<div class='property-card'>Error rendering property</div>"

    def _generate_error_email(self, error_msg: str) -> str:
        """Generate error email HTML"""
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .error {{ background: #ffcccc; padding: 20px; margin: 20px; }}
            </style>
        </head>
        <body>
            <div class="error">
                <h2>Error Generating Report</h2>
                <p>An error occurred while generating the daily report:</p>
                <p><strong>{error_msg}</strong></p>
            </div>
        </body>
        </html>
        """
